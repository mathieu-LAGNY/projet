#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 26 17:22:40 2020

@author: mathieu
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 14 12:43:19 2020

@author: mathieu
"""

from openpyxl import Workbook
import itertools

"""Exportation en xls"""

#.............................................................................
#E         result, la liste de tout les résultats
#          name, le nom du fichier à exporter
#Action    Permet d'exporter les résultats sur un graphes en format png
#S         none
def export_xls(result,name):
    
    wb = Workbook()
    ws = wb.active
        
    #En-têtes
    ws['A1'] = 'Masses'
    ws['B1'] = 'Probabilités'
    
    #Valeurs
    for i in range(len(result)):
       ws['A'+str(i+2)] = result[i][0]
       ws['B'+str(i+2)] = result[i][1]
    
    try :
        wb.save('../'+name+'.xlsx')
    except :
        wb.save(name+'.xlsx')
#.............................................................................
"""
#.............................................................................
#E         none
#Action    Permet l'importation d'un fichier xls pour le calcul inverse
#S         none   
def import_xls():
    
    try :
        
        wb = xlrd.open_workbook('fichier/'+Eimp.get()+'.xls')
        sh = wb.sheet_by_name(wb.sheet_names()[0])
        
        result = []
        result.append(sh["A"])
        result.append(sh["B"])
        
        Lalerti = Label(F3i,text='Le document est en traitement',fg='green',padx=20) 
        Lalerti.grid(row=1,column=1)
        
    except :
        
        Lalerti = Label(F3i,text='Le document est incorrect',fg='red',padx=20) 
        Lalerti.grid(row=1,column=1)
#............................................................................."""